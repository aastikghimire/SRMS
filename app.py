from __future__ import annotations

import base64
import datetime as dt
import io
import json
import math
import os
import re
import subprocess
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import plotly.express as px
import streamlit as st


# -----------------------------
# Page + small styling helpers
# -----------------------------

st.set_page_config(
    page_title="SRMS • School Result Management System",
    page_icon="📘",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
/* Dark-mode friendly tweaks (Streamlit theme still controls most colors) */
.srms-muted { opacity: 0.85; }
.srms-card {
  border: 1px solid rgba(255,255,255,0.10);
  border-radius: 14px;
  padding: 14px 16px;
  background: rgba(255,255,255,0.03);
}
.srms-small { font-size: 0.92rem; }
.srms-table { border-collapse: collapse; width: 100%; }
.srms-table th, .srms-table td { padding: 8px 10px; border-bottom: 1px solid rgba(255,255,255,0.10); }
.srms-right { text-align: right; }
.srms-center { text-align: center; }
</style>
""",
    unsafe_allow_html=True,
)


# -----------------------------
# Configuration models
# -----------------------------


def _slug(s: str) -> str:
    s = str(s).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "subject"


@dataclass(frozen=True)
class SubjectConfig:
    name: str
    theory_credit_hours: float
    practical_credit_hours: float
    theory_max: float
    practical_max: float

    @property
    def key(self) -> str:
        return _slug(self.name)

    @property
    def total_max(self) -> float:
        return float(self.theory_max) + float(self.practical_max)

    @property
    def total_credit_hours(self) -> float:
        return float(self.theory_credit_hours) + float(self.practical_credit_hours)


@dataclass(frozen=True)
class GradeBand:
    min_percent: float
    max_percent: float
    grade: str
    grade_point: float


def default_subjects() -> List[SubjectConfig]:
    return [
        # Demo split requested:
        # - Core subjects: 3.75 (TH) / 1.25 (PR)
        # - Social/Optional: 3.0 (TH) / 1.0 (PR)
        SubjectConfig("English", 3.75, 1.25, 75, 25),
        SubjectConfig("Nepali", 3.75, 1.25, 75, 25),
        SubjectConfig("Mathematics", 3.75, 1.25, 75, 25),
        SubjectConfig("Science", 3.75, 1.25, 75, 25),
        SubjectConfig("Social Studies", 3.0, 1.0, 75, 25),
    ]


def default_grade_scale() -> List[GradeBand]:
    # NEB-style percent-to-grade mapping (schools can edit)
    return [
        GradeBand(90, 100, "A+", 4.0),
        GradeBand(80, 90, "A", 3.6),
        GradeBand(70, 80, "B+", 3.2),
        GradeBand(60, 70, "B", 2.8),
        GradeBand(50, 60, "C+", 2.4),
        GradeBand(40, 50, "C", 2.0),
        GradeBand(35, 40, "D", 1.6),
        GradeBand(0, 35, "NG", 0.0),
    ]


def default_cgpa_scale() -> List[Dict[str, Any]]:
    # Grade mapping for final CGPA/average GP (schools can edit)
    return [
        {"min_cgpa": 3.61, "max_cgpa": 4.0, "grade": "A+"},
        {"min_cgpa": 3.21, "max_cgpa": 3.60, "grade": "A"},
        {"min_cgpa": 2.81, "max_cgpa": 3.20, "grade": "B+"},
        {"min_cgpa": 2.41, "max_cgpa": 2.80, "grade": "B"},
        {"min_cgpa": 2.01, "max_cgpa": 2.40, "grade": "C+"},
        {"min_cgpa": 1.61, "max_cgpa": 2.00, "grade": "C"},
        {"min_cgpa": 1.21, "max_cgpa": 1.60, "grade": "D"},
        {"min_cgpa": 0.0, "max_cgpa": 1.20, "grade": "NG"},
    ]


def _migrate_default_grading_config(cfg: Dict[str, Any]) -> None:
    grade_scale = cfg.get("grade_scale") or []
    grade_labels = [getattr(b, "grade", "") for b in grade_scale]
    if grade_labels == ["A+", "A", "B+", "B", "C+", "C", "D+", "D", "E"]:
        cfg["grade_scale"] = default_grade_scale()

    cgpa_scale = cfg.get("cgpa_scale") or []
    cgpa_labels = [str(r.get("grade", "")) for r in cgpa_scale if isinstance(r, dict)]
    if cgpa_labels == ["A+", "A", "B+", "B", "C+", "C", "D+", "D", "E"]:
        cfg["cgpa_scale"] = default_cgpa_scale()


def _ensure_state() -> None:
    if "auth" not in st.session_state:
        st.session_state.auth = {"logged_in": False, "username": ""}

    if "srms_config" not in st.session_state:
        st.session_state.srms_config = {
            "school_id": "",
            "school_name": "Your School Name",
            "school_address": "Municipality, District, Nepal",
            "exam_name": "Secondary Education Examination (SEE)",
            "academic_year": f"{dt.date.today().year}/{dt.date.today().year + 1}",
            "logo_data_url": "",
            "pass_percent": 35.0,
            "subjects": default_subjects(),
            "grade_scale": default_grade_scale(),
            "cgpa_scale": default_cgpa_scale(),
        }

    if "srms_data" not in st.session_state:
        st.session_state.srms_data = {
            "raw_df": None,
            "clean_df": None,
            "ledger_df": None,
        }

    if "verification_links" not in st.session_state:
        st.session_state.verification_links = {}

    _migrate_default_grading_config(st.session_state.srms_config)


_ensure_state()


def reset_runtime_school_data() -> None:
    st.session_state.srms_data = {"raw_df": None, "clean_df": None, "ledger_df": None}
    st.session_state.verification_links = {}


# -----------------------------
# Supabase + licensing helpers
# -----------------------------


def _secret_value(name: str, default: str = "") -> str:
    nested_names = {
        "SUPABASE_URL": ("supabase", "url"),
        "SUPABASE_SERVICE_ROLE_KEY": ("supabase", "service_role_key"),
        "SUPABASE_ANON_KEY": ("supabase", "anon_key"),
        "SUPABASE_STORAGE_BUCKET": ("supabase", "storage_bucket"),
        "SRMS_OWNER_ID": ("srms", "owner_id"),
        "SRMS_OWNER_PASSWORD": ("srms", "owner_password"),
    }
    try:
        if name in st.secrets:
            return str(st.secrets[name]).strip()
        group_key = nested_names.get(name)
        if group_key:
            group, key = group_key
            if group in st.secrets and key in st.secrets[group]:
                return str(st.secrets[group][key]).strip()
    except Exception:
        pass
    return str(os.environ.get(name, default) or "").strip()


def supabase_url() -> str:
    return _secret_value("SUPABASE_URL").rstrip("/")


def supabase_key() -> str:
    return _secret_value("SUPABASE_SERVICE_ROLE_KEY") or _secret_value("SUPABASE_ANON_KEY")


def supabase_storage_bucket() -> str:
    return _secret_value("SUPABASE_STORAGE_BUCKET", "marksheets") or "marksheets"


def supabase_configured() -> bool:
    return bool(supabase_url() and supabase_key())


def supabase_request(
    method: str,
    path: str,
    *,
    json_body: Optional[Any] = None,
    raw_body: Optional[bytes] = None,
    content_type: str = "application/json",
    extra_headers: Optional[Dict[str, str]] = None,
) -> Any:
    if not supabase_configured():
        raise RuntimeError("Supabase is not configured. Add SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY.")

    url = f"{supabase_url()}{path}"
    key = supabase_key()
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": content_type,
    }
    if extra_headers:
        headers.update(extra_headers)

    data: Optional[bytes] = None
    if json_body is not None:
        data = json.dumps(json_body).encode("utf-8")
    elif raw_body is not None:
        data = raw_body

    req = urllib.request.Request(url, data=data, headers=headers, method=method.upper())
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = resp.read()
            if not body:
                return None
            text = body.decode("utf-8")
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                return text
    except urllib.error.HTTPError as e:
        details = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Supabase request failed ({e.code}): {details}") from e
    except urllib.error.URLError as e:
        raise RuntimeError(f"Could not reach Supabase: {e.reason}") from e


def _parse_date(value: Any) -> Optional[dt.date]:
    if value is None or value == "":
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, dt.date):
        return value
    try:
        return dt.date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def _bool_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    try:
        if pd.isna(value):
            return False
    except Exception:
        pass
    return str(value).strip().lower() in {"1", "true", "yes", "y", "active"}


def _school_filter_value(value: str) -> str:
    return urllib.parse.quote(str(value).strip(), safe="")


def fetch_school_by_id(school_id: str) -> Optional[Dict[str, Any]]:
    school_id = str(school_id or "").strip()
    if not school_id:
        return None
    rows = supabase_request("GET", f"/rest/v1/schools?id=eq.{_school_filter_value(school_id)}&select=*")
    if isinstance(rows, list) and rows:
        return dict(rows[0])
    return None


def authenticate_school(school_id: str, password: str) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    try:
        school = fetch_school_by_id(school_id)
    except Exception as e:
        return None, str(e)
    if not school:
        return None, "School ID not found."
    if str(school.get("password", "")) != str(password):
        return None, "Invalid password."
    if not _bool_value(school.get("is_active", True)):
        return None, "This school account is inactive."
    expiry = _parse_date(school.get("expiry_date"))
    if expiry and expiry < dt.date.today():
        return None, f"This school's package expired on {expiry.isoformat()}."
    return school, None


def apply_school_profile(school: Dict[str, Any]) -> None:
    cfg = st.session_state.srms_config
    cfg["school_id"] = str(school.get("id", "") or "").strip()
    cfg["school_name"] = str(school.get("name", "") or cfg.get("school_name", "")).strip()
    if school.get("logo_url"):
        cfg["logo_data_url"] = str(school["logo_url"]).strip()


def owner_credentials() -> Tuple[str, str]:
    return _secret_value("SRMS_OWNER_ID"), _secret_value("SRMS_OWNER_PASSWORD")


def authenticate_owner(username: str, password: str) -> bool:
    owner_id, owner_password = owner_credentials()
    return bool(owner_id and owner_password and username == owner_id and password == owner_password)


def is_owner() -> bool:
    return st.session_state.auth.get("role") == "owner"


def fetch_schools() -> List[Dict[str, Any]]:
    rows = supabase_request("GET", "/rest/v1/schools?select=*&order=name.asc")
    return [dict(r) for r in rows] if isinstance(rows, list) else []


def upsert_schools(rows: List[Dict[str, Any]]) -> Any:
    payload: List[Dict[str, Any]] = []
    for row in rows:
        school_id = str(row.get("id", "") or "").strip()
        name = str(row.get("name", "") or "").strip()
        if not school_id or not name:
            continue
        expiry = _parse_date(row.get("expiry_date"))
        payload.append(
            {
                "id": school_id,
                "name": name,
                "password": str(row.get("password", "") or ""),
                "expiry_date": expiry.isoformat() if expiry else None,
                "logo_url": str(row.get("logo_url", "") or "").strip() or None,
                "is_active": _bool_value(row.get("is_active", True)),
            }
        )
    if not payload:
        raise RuntimeError("No valid school rows to save.")
    return supabase_request(
        "POST",
        "/rest/v1/schools?on_conflict=id",
        json_body=payload,
        extra_headers={"Prefer": "resolution=merge-duplicates,return=representation"},
    )


def _path_slug(value: Any, default: str = "item") -> str:
    slug = _slug(str(value or default))
    return slug[:80] or default


def marksheet_storage_path(student_row: pd.Series, cfg: Dict[str, Any]) -> str:
    school_id = _path_slug(cfg.get("school_id") or cfg.get("school_name") or "school")
    academic_year = _path_slug(cfg.get("academic_year") or dt.date.today().year)
    class_name = _path_slug(student_row.get("Class", "class"))
    roll = _path_slug(student_row.get("Roll_No", "roll"))
    symbol = _path_slug(student_row.get("Symbol_No", "symbol"))
    return f"{school_id}/{academic_year}/{class_name}/roll_{roll}_{symbol}.pdf"


def public_storage_url(path: str) -> str:
    encoded_path = urllib.parse.quote(path, safe="/")
    return f"{supabase_url()}/storage/v1/object/public/{supabase_storage_bucket()}/{encoded_path}"


def upload_pdf_to_storage(path: str, pdf_bytes: bytes) -> None:
    encoded_bucket = urllib.parse.quote(supabase_storage_bucket(), safe="")
    encoded_path = urllib.parse.quote(path, safe="/")
    supabase_request(
        "POST",
        f"/storage/v1/object/{encoded_bucket}/{encoded_path}",
        raw_body=pdf_bytes,
        content_type="application/pdf",
        extra_headers={"x-upsert": "true"},
    )


def record_marksheet_metadata(student_row: pd.Series, cfg: Dict[str, Any], path: str, public_url: str) -> Optional[str]:
    payload = {
        "school_id": str(cfg.get("school_id", "") or ""),
        "roll_no": str(student_row.get("Roll_No", "") or ""),
        "symbol_no": str(student_row.get("Symbol_No", "") or ""),
        "student_name": str(student_row.get("Student_Name", "") or ""),
        "academic_year": str(cfg.get("academic_year", "") or ""),
        "storage_path": path,
        "public_url": public_url,
        "created_at": dt.datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }
    try:
        supabase_request(
            "POST",
            "/rest/v1/marksheets?on_conflict=storage_path",
            json_body=payload,
            extra_headers={"Prefer": "resolution=merge-duplicates,return=minimal"},
        )
        return None
    except Exception as e:
        # The storage upload is the important part. Metadata is optional if the table is not created yet.
        return str(e)


# -----------------------------
# Authentication (sidebar)
# -----------------------------


def render_auth_sidebar() -> None:
    auth = st.session_state.auth

    st.sidebar.markdown("## Admin Access")
    st.sidebar.markdown(
        '<div class="srms-muted srms-small">Owner login controls packages. School login unlocks its own SRMS workspace.</div>',
        unsafe_allow_html=True,
    )

    if auth.get("logged_in"):
        role_label = "Main Admin" if auth.get("role") == "owner" else "School Admin"
        st.sidebar.success(f"{role_label}: {auth.get('username', '')}")
        if auth.get("school_name"):
            st.sidebar.caption(str(auth["school_name"]))
        if auth.get("expiry_date"):
            st.sidebar.caption(f"Package expiry: {auth['expiry_date']}")
        if st.sidebar.button("Logout", use_container_width=True):
            st.session_state.auth = {"logged_in": False, "username": ""}
            reset_runtime_school_data()
            st.rerun()
        st.sidebar.divider()
        return

    username = st.sidebar.text_input("School ID / Owner ID", value="", key="srms_login_username")
    password = st.sidebar.text_input("Password", value="", type="password", key="srms_login_password")

    if st.sidebar.button("Login", type="primary", use_container_width=True):
        if authenticate_owner(username, password):
            reset_runtime_school_data()
            st.session_state.auth = {"logged_in": True, "username": username, "role": "owner"}
            st.rerun()
        if not supabase_configured():
            st.sidebar.error("Supabase is not configured yet. Add SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY.")
        else:
            school, err = authenticate_school(username, password)
            if err:
                st.sidebar.error(err)
            elif school:
                reset_runtime_school_data()
                apply_school_profile(school)
                st.session_state.auth = {
                    "logged_in": True,
                    "username": str(school.get("id", "") or username),
                    "role": "school",
                    "school_id": str(school.get("id", "") or ""),
                    "school_name": str(school.get("name", "") or ""),
                    "expiry_date": str(school.get("expiry_date", "") or ""),
                }
                st.rerun()

    st.sidebar.divider()
    if not owner_credentials()[0]:
        st.sidebar.caption("Owner login is disabled until SRMS_OWNER_ID and SRMS_OWNER_PASSWORD are set.")
    if not supabase_configured():
        st.sidebar.caption("Supabase login is disabled until Supabase credentials are set.")


render_auth_sidebar()


def is_admin() -> bool:
    return bool(st.session_state.auth.get("logged_in"))


# -----------------------------
# Grade helpers
# -----------------------------


def normalize_grade_scale(bands: List[GradeBand]) -> List[GradeBand]:
    cleaned: List[GradeBand] = []
    for b in bands:
        mn = float(b.min_percent)
        mx = float(b.max_percent)
        if mx < mn:
            mn, mx = mx, mn
        cleaned.append(GradeBand(mn, mx, str(b.grade).strip(), float(b.grade_point)))
    cleaned.sort(key=lambda x: (-x.min_percent, -x.max_percent))
    return cleaned


def percent_to_grade(percent: float, scale: List[GradeBand]) -> Tuple[str, float]:
    if percent is None or (isinstance(percent, float) and math.isnan(percent)):
        return "NG", 0.0
    p = float(percent)
    if p < 0:
        return "NG", 0.0
    eps = 1e-9
    for band in scale:
        if p + eps >= band.min_percent and p <= band.max_percent + eps:
            return band.grade, float(band.grade_point)
    # fallback for out-of-range
    if p > max((b.max_percent for b in scale), default=100):
        top = max(scale, key=lambda x: x.max_percent)
        return top.grade, float(top.grade_point)
    return "NG", 0.0


def cgpa_to_final_grade(cgpa: float, cgpa_scale_rows: List[Dict[str, Any]]) -> str:
    if cgpa is None or (isinstance(cgpa, float) and math.isnan(cgpa)):
        return "NG"
    v = float(cgpa)
    rows = []
    for r in cgpa_scale_rows:
        try:
            rows.append(
                (float(r.get("min_cgpa", 0.0)), float(r.get("max_cgpa", 4.0)), str(r.get("grade", "")).strip())
            )
        except Exception:
            continue
    rows.sort(key=lambda x: (-x[0], -x[1]))
    eps = 1e-9
    for mn, mx, g in rows:
        if v + eps >= mn and v <= mx + eps:
            return g or "NG"
    return "NG"


# -----------------------------
# File template + parsing
# -----------------------------


def expected_columns_for_subjects(subjects: List[SubjectConfig]) -> List[str]:
    cols = ["Roll_No", "Student_Name", "Symbol_No", "DOB_BS", "Class", "Section"]
    for s in subjects:
        k = s.key
        cols.extend([f"{k}_TH", f"{k}_PR"])
    return cols


def build_upload_template(subjects: List[SubjectConfig]) -> pd.DataFrame:
    cols = expected_columns_for_subjects(subjects)
    df = pd.DataFrame(columns=cols)
    return df


def read_uploaded_file(uploaded) -> pd.DataFrame:
    name = getattr(uploaded, "name", "") or ""
    name_lower = name.lower()
    if name_lower.endswith(".csv"):
        return pd.read_csv(uploaded)
    if name_lower.endswith(".xlsx") or name_lower.endswith(".xls"):
        return pd.read_excel(uploaded)
    # streamlit may not preserve extension; try best-effort
    try:
        return pd.read_excel(uploaded)
    except Exception:
        uploaded.seek(0)
        return pd.read_csv(uploaded)


def coerce_number(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def clean_and_validate(
    df: pd.DataFrame,
    subjects: List[SubjectConfig],
    pass_percent: float,
) -> Tuple[pd.DataFrame, List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []

    if df is None or df.empty:
        return pd.DataFrame(), ["Uploaded file is empty."], warnings

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    required_base = ["Roll_No", "Student_Name"]
    for c in required_base:
        if c not in df.columns:
            errors.append(f"Missing required column: {c}")

    if errors:
        return pd.DataFrame(), errors, warnings

    df["Roll_No"] = df["Roll_No"].astype(str).str.strip()
    df["Student_Name"] = df["Student_Name"].astype(str).str.strip()
    df = df[df["Roll_No"].ne("")].copy()
    df = df[df["Student_Name"].ne("")].copy()

    # Ensure deterministic order for bulk outputs
    df["_roll_sort"] = pd.to_numeric(df["Roll_No"], errors="coerce")
    df["_roll_sort"] = df["_roll_sort"].fillna(10**9)
    df = df.sort_values(by=["_roll_sort", "Roll_No", "Student_Name"], ascending=True).drop(columns=["_roll_sort"])

    for s in subjects:
        k = s.key
        th_col = f"{k}_TH"
        pr_col = f"{k}_PR"
        total_col = k

        if th_col not in df.columns and pr_col not in df.columns and total_col in df.columns:
            # Allow a single total mark column if user provided that
            df[th_col] = df[total_col]
            df[pr_col] = 0
            warnings.append(
                f"Subject '{s.name}': used '{total_col}' as Theory and set Practical=0. "
                f"Recommended columns are '{th_col}' and '{pr_col}'."
            )

        if th_col not in df.columns:
            df[th_col] = None
            warnings.append(f"Subject '{s.name}': missing '{th_col}'. Filled with 0.")
        if pr_col not in df.columns:
            df[pr_col] = None
            warnings.append(f"Subject '{s.name}': missing '{pr_col}'. Filled with 0.")

        # Track missing (for NG)
        df[f"{k}__missing"] = df[[th_col, pr_col]].isna().any(axis=1).astype(int)

        df[th_col] = df[th_col].map(coerce_number).fillna(0.0)
        df[pr_col] = df[pr_col].map(coerce_number).fillna(0.0)

        # Clamp negatives to 0
        neg_th = (df[th_col] < 0).sum()
        neg_pr = (df[pr_col] < 0).sum()
        if neg_th or neg_pr:
            warnings.append(f"Subject '{s.name}': negative marks were set to 0.")
        df.loc[df[th_col] < 0, th_col] = 0.0
        df.loc[df[pr_col] < 0, pr_col] = 0.0

        # Clamp maxima
        over_th = (df[th_col] > s.theory_max).sum()
        over_pr = (df[pr_col] > s.practical_max).sum()
        if over_th:
            warnings.append(
                f"Subject '{s.name}': {int(over_th)} theory mark(s) exceeded max ({s.theory_max}) and were capped."
            )
        if over_pr:
            warnings.append(
                f"Subject '{s.name}': {int(over_pr)} practical mark(s) exceeded max ({s.practical_max}) and were capped."
            )
        df.loc[df[th_col] > s.theory_max, th_col] = float(s.theory_max)
        df.loc[df[pr_col] > s.practical_max, pr_col] = float(s.practical_max)

        # Optional basic sanity
        if s.total_max <= 0:
            errors.append(f"Invalid subject config for '{s.name}': total max must be > 0.")
        if s.total_credit_hours <= 0:
            errors.append(f"Invalid subject config for '{s.name}': credit hours must be > 0.")

        # Add derived total percent marker for validation/preview.
        obtained = df[th_col] + df[pr_col]
        percent = obtained / float(s.total_max) * 100.0
        df[f"{k}__percent"] = percent
        df[f"{k}__pass"] = (percent >= float(pass_percent)).astype(int)

    return df, errors, warnings


# -----------------------------
# Calculation + ledger
# -----------------------------


def build_ledger(
    clean_df: pd.DataFrame,
    subjects: List[SubjectConfig],
    grade_scale: List[GradeBand],
    pass_percent: float,
    cgpa_scale_rows: List[Dict[str, Any]],
) -> pd.DataFrame:
    df = clean_df.copy()
    scale = normalize_grade_scale(grade_scale)

    total_credits = float(sum(s.total_credit_hours for s in subjects)) if subjects else 0.0
    if total_credits <= 0:
        total_credits = 1.0

    for s in subjects:
        k = s.key
        th_col = f"{k}_TH"
        pr_col = f"{k}_PR"

        obtained = df[th_col] + df[pr_col]

        # Component grades (theory/practical) for marksheet presentation
        if float(s.theory_max) > 0:
            th_percent = (df[th_col] / float(s.theory_max) * 100.0).clip(lower=0.0, upper=100.0)
            th_grades = th_percent.map(lambda p: percent_to_grade(float(p), scale)[0])
            th_gps = th_percent.map(lambda p: percent_to_grade(float(p), scale)[1])
        else:
            th_percent = pd.Series([0.0] * len(df), index=df.index)
            th_grades = pd.Series(["-"] * len(df), index=df.index)
            th_gps = pd.Series([0.0] * len(df), index=df.index)

        if float(s.practical_max) > 0:
            pr_percent = (df[pr_col] / float(s.practical_max) * 100.0).clip(lower=0.0, upper=100.0)
            pr_grades = pr_percent.map(lambda p: percent_to_grade(float(p), scale)[0])
            pr_gps = pr_percent.map(lambda p: percent_to_grade(float(p), scale)[1])
        else:
            pr_percent = pd.Series([0.0] * len(df), index=df.index)
            pr_grades = pd.Series(["-"] * len(df), index=df.index)
            pr_gps = pd.Series([0.0] * len(df), index=df.index)

        # If any input missing for this subject row -> NG (not graded)
        ng_mask = df[f"{k}__missing"].astype(int) == 1
        theory_ng_mask = th_percent < float(pass_percent)
        subject_ng_mask = ng_mask | theory_ng_mask

        # NEB-style WGP: TH and PR are weighted separately by their own credit hours.
        th_wgp = th_gps.astype(float) * float(s.theory_credit_hours)
        pr_wgp = pr_gps.astype(float) * float(s.practical_credit_hours)
        subject_wgp = th_wgp + pr_wgp
        subject_gp = (subject_wgp / float(s.total_credit_hours)).round(2)
        subject_grades = subject_gp.map(lambda gp: cgpa_to_final_grade(float(gp), cgpa_scale_rows))

        subject_grades = subject_grades.where(~subject_ng_mask, other="NG")
        subject_gp = subject_gp.where(~subject_ng_mask, other=0.0)
        th_wgp = th_wgp.where(~subject_ng_mask, other=0.0)
        pr_wgp = pr_wgp.where(~subject_ng_mask, other=0.0)
        subject_wgp = subject_wgp.where(~subject_ng_mask, other=0.0)
        th_grades = th_grades.where(~ng_mask, other="NG")
        pr_grades = pr_grades.where(~ng_mask, other="NG")
        th_gps = th_gps.where(~ng_mask, other=0.0)
        pr_gps = pr_gps.where(~ng_mask, other=0.0)

        pass_mask = ~subject_ng_mask

        df[f"{k}__obtained"] = obtained.round(2)
        df[f"{k}__grade"] = subject_grades
        df[f"{k}__th_grade"] = th_grades
        df[f"{k}__pr_grade"] = pr_grades
        df[f"{k}__th_gp"] = th_gps.round(2)
        df[f"{k}__pr_gp"] = pr_gps.round(2)
        df[f"{k}__th_wgp"] = th_wgp.round(3)
        df[f"{k}__pr_wgp"] = pr_wgp.round(3)
        df[f"{k}__wgp"] = subject_wgp.round(3)
        df[f"{k}__gp"] = subject_gp.round(2)
        df[f"{k}__status"] = pass_mask.map(lambda x: "Pass" if bool(x) else "Fail")

    # Weighted totals
    gp_weighted_cols = []
    for s in subjects:
        k = s.key
        gp_weighted_cols.append(f"{k}__wgp")

    df["Total_Grade_Points"] = df[gp_weighted_cols].sum(axis=1).round(3) if gp_weighted_cols else 0.0
    df["CGPA"] = (df["Total_Grade_Points"] / total_credits).round(2)

    status_cols = [f"{s.key}__status" for s in subjects]
    missing_cols = [f"{s.key}__missing" for s in subjects]
    grade_cols = [f"{s.key}__grade" for s in subjects]
    any_missing = df[missing_cols].eq(1).any(axis=1) if missing_cols else False
    any_ng_grade = df[grade_cols].eq("NG").any(axis=1) if grade_cols else False
    any_ng = any_missing | any_ng_grade
    any_fail = df[status_cols].eq("Fail").any(axis=1) if status_cols else False

    # Strict NG + promotion rule:
    # - If ANY subject is NG OR Fail => FAILED
    # - Only if ALL subjects are Pass => PROMOTED
    df["Remarks"] = "PROMOTED"
    df.loc[any_ng | any_fail, "Remarks"] = "FAILED"

    df["Final_Grade"] = df["CGPA"].map(lambda v: cgpa_to_final_grade(float(v), cgpa_scale_rows))

    # If NG exists anywhere, show the calculated CGPA but keep the final grade as NG.
    df.loc[any_ng, "Final_Grade"] = "NG"

    # Friendly ledger view (stable column ordering)
    front = ["Roll_No", "Student_Name", "Symbol_No", "DOB_BS", "Class", "Section"]
    per_subject_cols: List[str] = []
    for s in subjects:
        k = s.key
        per_subject_cols.extend(
            [
                f"{k}_TH",
                f"{k}_PR",
                f"{k}__obtained",
                f"{k}__th_grade",
                f"{k}__pr_grade",
                f"{k}__grade",
                f"{k}__gp",
                f"{k}__status",
            ]
        )
    tail = ["Total_Grade_Points", "CGPA", "Final_Grade", "Remarks"]

    cols = [c for c in (front + per_subject_cols + tail) if c in df.columns]
    return df[cols].copy()


# -----------------------------
# Marksheet HTML + PDF
# -----------------------------


def _escape_html(s: Any) -> str:
    x = "" if s is None else str(s)
    return (
        x.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def image_bytes_to_data_url(img_bytes: bytes, mime: str) -> str:
    b64 = base64.b64encode(img_bytes).decode("ascii")
    return f"data:{mime};base64,{b64}"


def qr_code_data_url(payload: str) -> str:
    if not payload:
        return ""
    try:
        import qrcode  # type: ignore

        qr = qrcode.QRCode(version=None, box_size=6, border=1)
        qr.add_data(payload)
        qr.make(fit=True)
        img = qr.make_image(fill_color="#0b1220", back_color="white")
        out = io.BytesIO()
        img.save(out, format="PNG")
        return image_bytes_to_data_url(out.getvalue(), "image/png")
    except Exception:
        return ""


def grading_key_html(cgpa_scale_rows: List[Dict[str, Any]]) -> str:
    rows = []
    for r in cgpa_scale_rows or []:
        try:
            g = str(r.get("grade", "")).strip()
            mn = float(r.get("min_cgpa", 0.0))
            mx = float(r.get("max_cgpa", 4.0))
            if not g:
                continue
            rows.append((g, mn, mx))
        except Exception:
            continue
    rows.sort(key=lambda x: (-x[1], -x[2], x[0]))
    if not rows:
        return ""
    items = "".join(
        [
            f"""
            <div class="gk-item">
              <span class="gk-grade">{_escape_html(g)}</span>
              <span class="gk-range">{mn:.2f} - {mx:.2f}</span>
            </div>
            """
            for g, mn, mx in rows
        ]
    )
    return f"""
      <div class="grading-key">
        <div class="gk-title">Grading Key (CGPA)</div>
        <div class="gk-grid">{items}</div>
      </div>
    """


def marksheet_html(
    student_row: pd.Series,
    cfg: Dict[str, Any],
    subjects: List[SubjectConfig],
    verification_url: str = "",
) -> str:
    school_name_raw = str(cfg.get("school_name", "") or "").strip()
    school_name = _escape_html(school_name_raw)
    school_address = _escape_html(cfg.get("school_address", ""))
    exam_name = _escape_html(cfg.get("exam_name", ""))
    academic_year = _escape_html(cfg.get("academic_year", ""))

    roll_no = _escape_html(student_row.get("Roll_No", ""))
    student_name = _escape_html(student_row.get("Student_Name", ""))
    symbol_no = _escape_html(student_row.get("Symbol_No", ""))
    dob_bs = _escape_html(student_row.get("DOB_BS", ""))
    class_name = _escape_html(student_row.get("Class", ""))
    section = _escape_html(student_row.get("Section", ""))
    remarks = _escape_html(student_row.get("Remarks", ""))
    cgpa_val = student_row.get("CGPA", "")
    final_grade = _escape_html(student_row.get("Final_Grade", ""))
    cgpa_display = "NG"
    if cgpa_val is not None and not (isinstance(cgpa_val, float) and math.isnan(cgpa_val)):
        cgpa_display = _escape_html(cgpa_val)

    today = dt.date.today().strftime("%Y-%m-%d")
    logo_data_url = str(cfg.get("logo_data_url", "") or "").strip()
    logo_src = _escape_html(logo_data_url)
    verification_url = str(verification_url or "").strip()
    verification_url_safe = _escape_html(verification_url)
    qr_img = qr_code_data_url(verification_url)
    if verification_url and qr_img:
        qr_html = f"<img src='{qr_img}' alt='Verification QR'/>"
    elif verification_url:
        qr_html = f"<a href='{verification_url_safe}'>Verify PDF</a>"
    else:
        qr_html = "<b>QR</b><br/>Verification"
    watermark_style = ""
    if logo_data_url:
        watermark_style = f"background-image: url('{logo_src}');"

    subject_count = max(len(subjects), 1)
    extra_subjects = max(subject_count - 5, 0)
    table_font_px = max(8.4, 10.6 - (extra_subjects * 0.38))
    table_cell_y_px = max(1.8, 4.2 - (extra_subjects * 0.42))
    table_cell_x_px = max(4.2, 7.0 - (extra_subjects * 0.35))
    compact_gap_px = max(5.0, 9.0 - (extra_subjects * 0.65))
    school_name_len = len(school_name_raw)
    if school_name_len > 56:
        school_name_font_px = 16.0
    elif school_name_len > 42:
        school_name_font_px = 18.0
    elif school_name_len > 30:
        school_name_font_px = 20.0
    else:
        school_name_font_px = 23.0

    rows_html = ""
    for s in subjects:
        k = s.key
        th_grade = _escape_html(student_row.get(f"{k}__th_grade", ""))
        pr_grade = _escape_html(student_row.get(f"{k}__pr_grade", ""))
        grade = _escape_html(student_row.get(f"{k}__grade", ""))
        th_gp = student_row.get(f"{k}__th_gp", "")
        pr_gp = student_row.get(f"{k}__pr_gp", "")
        status = _escape_html(student_row.get(f"{k}__status", ""))
        # Split into TH and PR rows like the official sheet
        rows_html += f"""
          <tr>
            <td>{_escape_html(s.name)} Th</td>
            <td class="center">{_escape_html(s.theory_credit_hours)}</td>
            <td class="center">{th_grade}</td>
            <td class="center">{_escape_html(th_gp)}</td>
            <td class="center">{status}</td>
          </tr>
          <tr>
            <td>{_escape_html(s.name)} Pr</td>
            <td class="center">{_escape_html(s.practical_credit_hours)}</td>
            <td class="center">{pr_grade}</td>
            <td class="center">{_escape_html(pr_gp)}</td>
            <td class="center">{status}</td>
          </tr>
          <tr class="sub-total">
            <td><b>{_escape_html(s.name)} (Final)</b></td>
            <td class="center"><b>{_escape_html(s.total_credit_hours)}</b></td>
            <td class="center"><b>{grade}</b></td>
            <td class="center"><b>{_escape_html(student_row.get(f"{k}__gp",""))}</b></td>
            <td class="center"><b>{status}</b></td>
          </tr>
        """

    gk_html = grading_key_html(cfg.get("cgpa_scale", []))

    css = f"""
    <style>
      @page {{ size: A4; margin: 10mm; }}
      html, body {{ margin: 0; padding: 0; }}
      body {{
        font-family: Inter, Poppins, Arial, Helvetica, sans-serif;
        color: #0b1220;
        background: white;
      }}

      .marksheet {{
        --table-font: {table_font_px:.2f}px;
        --table-cell-y: {table_cell_y_px:.2f}px;
        --table-cell-x: {table_cell_x_px:.2f}px;
        --compact-gap: {compact_gap_px:.2f}px;
        --school-name-font: {school_name_font_px:.2f}px;
        width: 190mm;
        height: 277mm;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        gap: 7px;
        padding: 7mm 8mm;
        box-sizing: border-box;
        background: white;
        position: relative;
        margin: 0 auto;
        overflow: hidden;
        border: 3px double #1a2a6c;
      }}

      .marksheet::before {{
        content: "";
        position: absolute;
        inset: 16mm 12mm;
        {watermark_style}
        background-repeat: no-repeat;
        background-position: center;
        background-size: 62%;
        opacity: 0.075;
        pointer-events: none;
        z-index: 0;
      }}
      .layer {{ position: relative; z-index: 1; }}
      .main-layer {{
        flex: 1 1 auto;
        min-height: 0;
        overflow: hidden;
      }}
      .footer-layer {{
        flex: 0 0 auto;
        padding-top: 5px;
      }}

      .header {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 9px;
      }}

      .header .logo {{
        width: 64px;
        height: 64px;
        border: 0;
        border-radius: 0;
        background: transparent;
        display: flex;
        align-items: center;
        justify-content: center;
        overflow: visible;
        flex: 0 0 auto;
      }}
      .header .logo img {{
        width: 100%;
        height: 100%;
        display: block;
        object-fit: contain;
        mix-blend-mode: multiply;
      }}
      .header .logo .ph {{ font-size: 11px; color: #64748b; }}

      .header .school {{
        flex: 1 1 auto;
        text-align: center;
        min-width: 0;
      }}
      .header .school .name {{
        font-size: var(--school-name-font);
        line-height: 1.08;
        font-weight: 800;
        letter-spacing: 0;
        white-space: normal;
        overflow: visible;
        overflow-wrap: anywhere;
        margin: 0;
      }}
      .header .school .addr {{
        margin-top: 3px;
        font-size: 10.5px;
        font-weight: 500;
        color: #334155;
      }}
      .header .school .doc-title {{
        margin-top: 6px;
        font-size: 13.8px;
        line-height: 1;
        font-weight: 900;
        color: #1a2a6c;
        letter-spacing: 0;
        text-transform: uppercase;
      }}

      .header .qr {{
        width: 76px;
        height: 76px;
        border: 1px dashed rgba(26, 42, 108, 0.42);
        border-radius: 8px;
        display: flex;
        align-items: center;
        justify-content: center;
        text-align: center;
        font-size: 7px;
        color: #334155;
        flex: 0 0 auto;
        background: rgba(255, 255, 255, 0.76);
      }}
      .header .qr img {{
        width: 66px;
        height: 66px;
        display: block;
      }}
      .header .qr a {{
        color: #1a2a6c;
        font-weight: 800;
        text-decoration: none;
      }}

      .subheader {{
        margin-top: var(--compact-gap);
        display: flex;
        justify-content: space-between;
        gap: 8px;
        font-size: 9.6px;
        color: #0f172a;
        border-top: 1px solid rgba(148, 163, 184, 0.32);
        border-bottom: 1px solid rgba(148, 163, 184, 0.32);
        padding: 5px 0;
      }}

      .info {{
        margin-top: var(--compact-gap);
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 4px 12px;
        font-size: 9.6px;
      }}
      .info div:nth-child(even) {{
        text-align: right;
      }}

      .marks-table {{
        width: 100%;
        border-collapse: collapse;
        table-layout: fixed;
        font-size: var(--table-font);
        line-height: 1.16;
        margin-top: var(--compact-gap);
      }}
      .marks-table th,
      .marks-table td {{
        border: 1px solid rgba(203, 213, 225, 0.72);
        padding: var(--table-cell-y) var(--table-cell-x);
        vertical-align: middle;
      }}
      .marks-table th {{
        background: #1a2a6c;
        color: #e2e8f0;
        font-weight: 800;
        text-align: left;
      }}
      .marks-table tbody tr:nth-child(even) td {{ background: rgba(248, 250, 252, 0.84); }}
      .marks-table tbody tr:nth-child(odd) td {{ background: rgba(255, 255, 255, 0.84); }}
      .center {{ text-align: center; }}
      .marks-table .sub-total td {{
        background: rgba(226, 232, 255, 0.88);
        font-weight: 800;
      }}

      .summary {{
        margin-top: var(--compact-gap);
        display: flex;
        justify-content: space-between;
        align-items: center;
        gap: 8px;
        font-size: 9.6px;
      }}
      .pill {{
        display: inline-block;
        padding: 4px 8px;
        border-radius: 999px;
        border: 1px solid rgba(26, 42, 108, 0.16);
        background: rgba(255, 255, 255, 0.72);
      }}

      .grading-key {{
        margin-top: var(--compact-gap);
        border: 1px solid rgba(26, 42, 108, 0.20);
        border-radius: 8px;
        padding: 7px;
        box-sizing: border-box;
        width: 42%;
        min-width: 205px;
        background: linear-gradient(180deg, rgba(248, 250, 252, 0.86), rgba(255, 255, 255, 0.74));
      }}
      .gk-title {{
        font-weight: 800;
        color: #1a2a6c;
        margin-bottom: 5px;
        font-size: 9.4px;
        letter-spacing: 0;
        text-transform: uppercase;
      }}
      .gk-grid {{
        display: flex;
        flex-direction: column;
        gap: 3px;
      }}
      .gk-item {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 8px;
        min-width: 0;
        padding: 3px 6px;
        border: 1px solid rgba(148, 163, 184, 0.38);
        border-radius: 6px;
        background: rgba(255, 255, 255, 0.70);
        font-size: 8.6px;
      }}
      .gk-grade {{
        min-width: 21px;
        height: 14px;
        border-radius: 999px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        color: white;
        background: #1a2a6c;
        font-weight: 800;
      }}
      .gk-range {{
        color: #334155;
        font-weight: 700;
        white-space: nowrap;
      }}

      .signatures {{
        display: flex;
        justify-content: space-between;
        gap: 10px;
        align-items: end;
      }}
      .signatures .sig {{
        flex: 1 1 0;
        min-width: 0;
        height: 31px;
        border-top: 1px solid #0f172a;
        padding-top: 5px;
        text-align: center;
        font-size: 9.4px;
        color: #0f172a;
        display: flex;
        align-items: center;
        justify-content: center;
      }}

      .print-btn {{
        position: absolute;
        right: 8mm;
        bottom: 7mm;
        border: 1px solid #1d4ed8;
        background: #eff6ff;
        color: #0f172a;
        padding: 5px 8px;
        border-radius: 8px;
        font-size: 10px;
        cursor: pointer;
        z-index: 2;
      }}

      .page-break {{
        break-after: page;
        page-break-after: always;
        height: 0;
        margin: 0;
        padding: 0;
      }}

      @media screen {{
        body {{ padding: 10px 0; background: #e5e7eb; }}
        .marksheet {{ box-shadow: 0 10px 30px rgba(15, 23, 42, 0.16); }}
      }}

      @media print {{
        .print-btn {{ display: none !important; }}
        html, body {{
          background: white !important;
          overflow: hidden;
        }}
        body {{ padding: 0 !important; }}
        .marksheet {{
          width: 190mm;
          height: 277mm;
          margin: 0 auto;
          box-shadow: none !important;
          break-inside: avoid;
          page-break-inside: avoid;
        }}
        * {{
          -webkit-print-color-adjust: exact !important;
          print-color-adjust: exact !important;
        }}
      }}
    </style>
    """

    html = f"""
    <html>
      <head>{css}</head>
      <body>
        <div class="marksheet">
          <button class="print-btn" onclick="window.print()">Print</button>

          <div class="layer main-layer">
            <div class="header">
              <div class="logo">
                {"<img src='" + logo_src + "' alt='Logo'/>" if logo_src else "<div class='ph'>LOGO</div>"}
              </div>
              <div class="school">
                <div class="name">{school_name}</div>
                <div class="addr">{school_address}</div>
                <div class="doc-title"> GRADESHEET </div>
              </div>
              <div class="qr">{qr_html}</div>
            </div>

            <div class="subheader">
              <div><b>{exam_name}</b></div>
              <div><b>Academic Year:</b> {academic_year}</div>
            </div>

            <div class="info">
              <div><b>Student Name:</b> {student_name}</div>
              <div><b>Roll No:</b> {roll_no}</div>
              <div><b>Student Symbol No:</b> {symbol_no}</div>
              <div><b>Date of Birth (BS):</b> {dob_bs}</div>
              <div><b>Class:</b> {class_name}</div>
              <div><b>Section:</b> {section}</div>
            </div>

            <table class="marks-table">
              <colgroup>
                <col style="width: 29%;">
                <col style="width: 25%;">
                <col style="width: 15%;">
                <col style="width: 19%;">
                <col style="width: 12%;">
              </colgroup>
              <thead>
                <tr>
                  <th>Subject</th>
                  <th class="center">Credit Hours (Cr)</th>
                  <th class="center">Grade</th>
                  <th class="center">Grade Point</th>
                  <th class="center">Status</th>
                </tr>
              </thead>
              <tbody>
                {rows_html}
              </tbody>
            </table>

            <div class="summary">
              <div>
                <span class="pill"><b>CGPA:</b> {cgpa_display}</span>
                &nbsp;
                <span class="pill"><b>Final Grade:</b> {final_grade}</span>
              </div>
              <div>
                <span class="pill"><b>Remarks:</b> {remarks}</span>
              </div>
            </div>

            {gk_html}
          </div>

          <div class="layer footer-layer">
            <div class="signatures">
              <div class="sig">Class Teacher</div>
              <div class="sig">Exam Coordinator</div>
              <div class="sig">School Seal</div>
              <div class="sig">Principal</div>
            </div>
          </div>
        </div>
      </body>
    </html>
    """
    return html


def html_to_pdf_bytes(html: str) -> Tuple[Optional[bytes], Optional[str]]:
    """
    Converts HTML to PDF.
    Prefers xhtml2pdf if installed, then uses local Chrome/Edge headless printing.
    """
    # xhtml2pdf
    xhtml2pdf_error = ""
    try:
        from xhtml2pdf import pisa  # type: ignore

        out = io.BytesIO()
        res = pisa.CreatePDF(io.StringIO(html), dest=out, encoding="utf-8")
        if res.err:
            return None, "PDF generation failed (xhtml2pdf error)."
        return out.getvalue(), None
    except Exception as e:
        xhtml2pdf_error = str(e)

    chrome_error = ""
    browser_candidates = [
        os.environ.get("SRMS_CHROME_PATH", ""),
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    ]
    browser_path = next((p for p in browser_candidates if p and os.path.exists(p)), "")
    if browser_path:
        try:
            with tempfile.TemporaryDirectory(prefix="srms_pdf_") as tmp_dir:
                html_path = os.path.join(tmp_dir, "marksheet.html")
                pdf_path = os.path.join(tmp_dir, "marksheet.pdf")
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(html)
                cmd = [
                    browser_path,
                    "--headless=new",
                    "--disable-gpu",
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--print-to-pdf-no-header",
                    f"--print-to-pdf={pdf_path}",
                    "file:///" + html_path.replace("\\", "/"),
                ]
                res = subprocess.run(cmd, capture_output=True, text=True, timeout=45)
                if os.path.exists(pdf_path):
                    with open(pdf_path, "rb") as f:
                        pdf = f.read()
                    if pdf:
                        return pdf, None
                chrome_error = (res.stderr or res.stdout or f"Chrome exited with code {res.returncode}").strip()
        except Exception as e:
            chrome_error = str(e)
    else:
        chrome_error = "Chrome/Edge executable not found."

    # pdfkit fallback
    pdfkit_error = ""
    try:
        import pdfkit  # type: ignore

        options = {
            "page-size": "A4",
            "margin-top": "14mm",
            "margin-right": "14mm",
            "margin-bottom": "14mm",
            "margin-left": "14mm",
            "encoding": "UTF-8",
            "quiet": "",
        }
        pdf = pdfkit.from_string(html, False, options=options)
        if not pdf:
            return None, "PDF generation failed (pdfkit returned empty output)."
        return pdf, None
    except Exception as e:
        pdfkit_error = str(e)
        return None, (
            "PDF engine not available. Install Chrome/Edge or set SRMS_CHROME_PATH to your browser executable. "
            f"xhtml2pdf error: {xhtml2pdf_error or 'not available'}; "
            f"Chrome/Edge error: {chrome_error or 'not available'}; "
            f"pdfkit error: {pdfkit_error or 'not available'}"
        )


def verification_key(student_row: pd.Series) -> str:
    return str(student_row.get("Roll_No", "") or "").strip()


def verification_url_for_row(student_row: pd.Series, links: Optional[Dict[str, str]] = None) -> str:
    if not links:
        return ""
    return str(links.get(verification_key(student_row), "") or "")


def publish_marksheet_pdf(student_row: pd.Series, cfg: Dict[str, Any], subjects: List[SubjectConfig]) -> Tuple[Optional[str], Optional[str]]:
    if not supabase_configured():
        return None, "Supabase is not configured. Add SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY."
    path = marksheet_storage_path(student_row, cfg)
    public_url = public_storage_url(path)
    html = marksheet_html(student_row, cfg, subjects, verification_url=public_url)
    pdf_bytes, err = html_to_pdf_bytes(html)
    if err or not pdf_bytes:
        return None, err or "PDF generation returned empty output."
    try:
        upload_pdf_to_storage(path, pdf_bytes)
        metadata_warning = record_marksheet_metadata(student_row, cfg, path, public_url)
    except Exception as e:
        return None, str(e)
    if metadata_warning:
        st.warning(f"PDF uploaded, but metadata was not saved: {metadata_warning}")
    return public_url, None


def bulk_marksheet_pdf(
    ledger_df: pd.DataFrame,
    cfg: Dict[str, Any],
    subjects: List[SubjectConfig],
    verification_links: Optional[Dict[str, str]] = None,
) -> Tuple[Optional[bytes], Optional[str]]:
    # One PDF containing individual pages for each student, sorted by roll (ledger already sorted)
    pages: List[str] = []
    for _, row in ledger_df.iterrows():
        pages.append(marksheet_html(row, cfg, subjects, verification_url_for_row(row, verification_links)))

    # Stitch pages: keep head/body per page simple; for xhtml2pdf, page-breaks work if we use a single HTML doc.
    first = pages[0] if pages else ""
    m_style = re.search(r"(<style>.*?</style>)", first, flags=re.IGNORECASE | re.DOTALL)
    style_block = m_style.group(1) if m_style else ""

    stitched = f"""
    <html><head><meta charset="utf-8">{style_block}</head><body>
    """
    for i, page in enumerate(pages):
        # Extract body content from each page
        m = re.search(r"<body>(.*)</body>", page, flags=re.IGNORECASE | re.DOTALL)
        body = m.group(1) if m else page
        stitched += body
        if i < len(pages) - 1:
            stitched += '<div class="page-break"></div>'
    stitched += "</body></html>"

    return html_to_pdf_bytes(stitched)


def bulk_marksheet_print_html(
    ledger_df: pd.DataFrame,
    cfg: Dict[str, Any],
    subjects: List[SubjectConfig],
    verification_links: Optional[Dict[str, str]] = None,
) -> str:
    """
    Generates one long HTML file containing every student's marksheet separated by page breaks.
    This is intended for browser printing (Ctrl+P) as a single document.
    """
    pages: List[str] = []
    for _, row in ledger_df.iterrows():
        pages.append(marksheet_html(row, cfg, subjects, verification_url_for_row(row, verification_links)))

    first = pages[0] if pages else ""
    m_style = re.search(r"(<style>.*?</style>)", first, flags=re.IGNORECASE | re.DOTALL)
    style_block = m_style.group(1) if m_style else ""

    stitched = f"<html><head><meta charset='utf-8'>{style_block}</head><body>"
    for i, page in enumerate(pages):
        m = re.search(r"<body>(.*)</body>", page, flags=re.IGNORECASE | re.DOTALL)
        body = m.group(1) if m else page
        stitched += body
        if i < len(pages) - 1:
            stitched += "<div class='page-break'></div>"
    stitched += "</body></html>"
    return stitched


# -----------------------------
# Excel export (formatted)
# -----------------------------


def ledger_to_excel_bytes(ledger_df: pd.DataFrame, subjects: List[SubjectConfig]) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        sheet = "Ledger"
        df = ledger_df.copy()
        df.to_excel(writer, sheet_name=sheet, index=False)

        ws = writer.sheets[sheet]
        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill("solid", fgColor="1F2937")  # slate-ish
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, cell in enumerate(ws[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for c in col_cells[:2000]:
                v = "" if c.value is None else str(c.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

        # Center some columns
        center_cols = {"Roll_No", "CGPA", "Final_Grade", "Remarks", "Total_Grade_Points"}
        for s in subjects:
            k = s.key
            center_cols.update({f"{k}__grade", f"{k}__gp", f"{k}__status"})
        header_row = [c.value for c in ws[1]]
        for idx, name in enumerate(header_row, start=1):
            if name in center_cols:
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=idx).alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

    return out.getvalue()


# -----------------------------
# UI sections
# -----------------------------


st.title("School Result Management System (SRMS)")
st.caption("SEE-style grading • configurable subjects/credits • marksheet printing • bulk PDF • insights")

cfg = st.session_state.srms_config


tab_work, tab_print, tab_insights = st.tabs(["📥 Data & Ledger", "🖨️ Print Engine", "📊 Insights"])


def render_owner_school_accounts() -> None:
    if not is_owner():
        return
    st.sidebar.markdown("## School Packages")
    if not supabase_configured():
        st.sidebar.warning("Set Supabase credentials to manage school accounts.")
        return
    try:
        schools = fetch_schools()
    except Exception as e:
        st.sidebar.error(f"Could not load schools: {e}")
        return

    school_df = pd.DataFrame(schools, columns=["id", "name", "password", "expiry_date", "logo_url", "is_active"])
    if school_df.empty:
        school_df = pd.DataFrame(columns=["id", "name", "password", "expiry_date", "logo_url", "is_active"])
    school_df["expiry_date"] = pd.to_datetime(school_df["expiry_date"], errors="coerce")
    school_df["is_active"] = school_df["is_active"].fillna(True).map(_bool_value)
    edited_schools = st.sidebar.data_editor(
        school_df,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "id": st.column_config.TextColumn("School ID", required=True),
            "name": st.column_config.TextColumn("School Name", required=True),
            "password": st.column_config.TextColumn("Password", required=True),
            "expiry_date": st.column_config.DateColumn("Expiry Date", required=True),
            "logo_url": st.column_config.TextColumn("Logo URL"),
            "is_active": st.column_config.CheckboxColumn("Active", default=True),
        },
        key="srms_school_accounts_editor",
    )
    if st.sidebar.button("Save School Packages", type="primary", use_container_width=True):
        try:
            rows = edited_schools.to_dict(orient="records")
            upsert_schools(rows)
            st.sidebar.success("School packages saved.")
        except Exception as e:
            st.sidebar.error(f"Could not save schools: {e}")


def render_configuration() -> None:
    st.sidebar.markdown("## Configuration")
    if not is_admin():
        st.sidebar.info("Login as Admin to edit configuration.")
        return

    render_owner_school_accounts()

    with st.sidebar.expander("School & Exam", expanded=True):
        cfg["school_name"] = st.text_input("School Name", value=cfg["school_name"], disabled=not is_owner())
        cfg["school_address"] = st.text_input("School Address", value=cfg["school_address"])
        cfg["exam_name"] = st.text_input("Exam Name", value=cfg["exam_name"])
        cfg["academic_year"] = st.text_input("Academic Year", value=cfg["academic_year"])
        logo_file = st.file_uploader(
            "School Logo (PNG/JPG) • used in header + watermark",
            type=["png", "jpg", "jpeg", "webp"],
            accept_multiple_files=False,
        )
        if logo_file is not None:
            try:
                mime = getattr(logo_file, "type", "") or "image/png"
                cfg["logo_data_url"] = image_bytes_to_data_url(logo_file.getvalue(), mime)
                st.success("Logo loaded for marksheet.")
            except Exception as e:
                st.error(f"Could not load logo: {e}")
        if cfg.get("logo_data_url"):
            if st.button("Clear Logo", use_container_width=True):
                cfg["logo_data_url"] = ""
                st.rerun()

    with st.sidebar.expander("Account Source", expanded=False):
        st.caption("Owner credentials come from SRMS_OWNER_ID / SRMS_OWNER_PASSWORD.")
        st.caption("School credentials come from the Supabase `schools` table.")
        if st.session_state.auth.get("school_id"):
            st.info(f"Current school ID: {st.session_state.auth['school_id']}")

    with st.sidebar.expander("Subjects • Credits • Max Marks", expanded=True):
        subj_rows = [
            {
                "Subject": s.name,
                "Theory_CrHrs": float(s.theory_credit_hours),
                "Practical_CrHrs": float(s.practical_credit_hours),
                "Theory_Max": float(s.theory_max),
                "Practical_Max": float(s.practical_max),
            }
            for s in cfg["subjects"]
        ]
        subj_df = pd.DataFrame(subj_rows) if subj_rows else pd.DataFrame(
            columns=["Subject", "Theory_CrHrs", "Practical_CrHrs", "Theory_Max", "Practical_Max"]
        )
        edited = st.data_editor(
            subj_df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "Subject": st.column_config.TextColumn(required=True),
                "Theory_CrHrs": st.column_config.NumberColumn(min_value=0.0, step=0.25, required=True),
                "Practical_CrHrs": st.column_config.NumberColumn(min_value=0.0, step=0.25, required=True),
                "Theory_Max": st.column_config.NumberColumn(min_value=0.0, step=1.0, required=True),
                "Practical_Max": st.column_config.NumberColumn(min_value=0.0, step=1.0, required=True),
            },
        )

        new_subjects: List[SubjectConfig] = []
        for _, r in edited.dropna(how="all").iterrows():
            name = str(r.get("Subject", "")).strip()
            if not name:
                continue
            new_subjects.append(
                SubjectConfig(
                    name=name,
                    theory_credit_hours=float(r.get("Theory_CrHrs", 0) or 0),
                    practical_credit_hours=float(r.get("Practical_CrHrs", 0) or 0),
                    theory_max=float(r.get("Theory_Max", 0) or 0),
                    practical_max=float(r.get("Practical_Max", 0) or 0),
                )
            )
        cfg["subjects"] = new_subjects

        cfg["pass_percent"] = st.number_input(
            "Theory Pass Threshold (%)",
            min_value=0.0,
            max_value=100.0,
            value=float(cfg.get("pass_percent", 35.0)),
            step=1.0,
        )

    with st.sidebar.expander("Grading Scale (Percent → Grade / GP)", expanded=True):
        gs_rows = [
            {"Min_%": b.min_percent, "Max_%": b.max_percent, "Grade": b.grade, "Grade_Point": b.grade_point}
            for b in cfg["grade_scale"]
        ]
        gs_df = pd.DataFrame(gs_rows) if gs_rows else pd.DataFrame(columns=["Min_%", "Max_%", "Grade", "Grade_Point"])
        edited_gs = st.data_editor(
            gs_df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "Min_%": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=1.0, required=True),
                "Max_%": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=1.0, required=True),
                "Grade": st.column_config.TextColumn(required=True),
                "Grade_Point": st.column_config.NumberColumn(min_value=0.0, max_value=4.0, step=0.1, required=True),
            },
        )
        new_scale: List[GradeBand] = []
        for _, r in edited_gs.dropna(how="all").iterrows():
            g = str(r.get("Grade", "")).strip()
            if not g:
                continue
            new_scale.append(
                GradeBand(
                    min_percent=float(r.get("Min_%", 0) or 0),
                    max_percent=float(r.get("Max_%", 0) or 0),
                    grade=g,
                    grade_point=float(r.get("Grade_Point", 0) or 0),
                )
            )
        cfg["grade_scale"] = new_scale

    with st.sidebar.expander("Final Grade (CGPA → Grade)", expanded=False):
        cg_df = pd.DataFrame(cfg["cgpa_scale"])
        edited_cg = st.data_editor(
            cg_df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "min_cgpa": st.column_config.NumberColumn(min_value=0.0, max_value=4.0, step=0.01, required=True),
                "max_cgpa": st.column_config.NumberColumn(min_value=0.0, max_value=4.0, step=0.01, required=True),
                "grade": st.column_config.TextColumn(required=True),
            },
        )
        cfg["cgpa_scale"] = edited_cg.fillna("").to_dict(orient="records")


render_configuration()


with tab_work:
    left, right = st.columns([1.2, 1.0], gap="large")

    with left:
        st.markdown("### Upload & Process")
        st.markdown('<div class="srms-muted srms-small">Upload a CSV/XLSX with Roll_No, Student_Name and subject marks.</div>', unsafe_allow_html=True)

        subjects: List[SubjectConfig] = cfg["subjects"]
        if not subjects:
            st.error("No subjects configured. Add subjects in the sidebar (Admin).")
        else:
            temp_df = build_upload_template(subjects)
            csv_bytes = temp_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "Download Upload Template (CSV)",
                data=csv_bytes,
                file_name="srms_upload_template.csv",
                mime="text/csv",
                use_container_width=True,
            )

        uploaded = st.file_uploader(
            "Upload Marks File (CSV/XLSX)",
            type=["csv", "xlsx", "xls"],
            disabled=not is_admin(),
            help="Use columns like english_TH, english_PR (based on subject names).",
        )

        if uploaded is not None and is_admin():
            try:
                raw_df = read_uploaded_file(uploaded)
                st.session_state.srms_data["raw_df"] = raw_df
                st.success(f"Loaded {len(raw_df):,} row(s) from {uploaded.name}")
            except Exception as e:
                st.error(f"Could not read file: {e}")

        raw_df = st.session_state.srms_data.get("raw_df")
        if raw_df is not None:
            st.markdown("### Preview (Raw)")
            st.dataframe(raw_df.head(50), use_container_width=True)

        if st.button("Clean + Calculate Ledger", type="primary", disabled=not (is_admin() and raw_df is not None), use_container_width=True):
            clean_df, errs, warns = clean_and_validate(raw_df, subjects, float(cfg["pass_percent"]))
            if errs:
                for e in errs:
                    st.error(e)
            else:
                if warns:
                    with st.expander("Warnings", expanded=False):
                        for w in warns[:200]:
                            st.warning(w)
                        if len(warns) > 200:
                            st.info(f"{len(warns) - 200} more warning(s) hidden.")

                ledger_df = build_ledger(
                    clean_df,
                    subjects,
                    cfg["grade_scale"],
                    float(cfg["pass_percent"]),
                    cfg["cgpa_scale"],
                )
                st.session_state.srms_data["clean_df"] = clean_df
                st.session_state.srms_data["ledger_df"] = ledger_df
                st.success("Ledger calculated.")

    with right:
        st.markdown("### Ledger")
        ledger_df = st.session_state.srms_data.get("ledger_df")
        if ledger_df is None:
            st.info("After uploading, click **Clean + Calculate Ledger**.")
        else:
            st.dataframe(ledger_df, use_container_width=True, height=520)

            c1, c2 = st.columns([1, 1], gap="medium")
            with c1:
                st.metric("Students", f"{len(ledger_df):,}")
            with c2:
                promoted_rate = (ledger_df["Remarks"] == "PROMOTED").mean() * 100.0 if len(ledger_df) else 0.0
                st.metric("Promoted Rate", f"{promoted_rate:.1f}%")

            st.markdown("### Export Ledger (Excel)")
            try:
                excel_bytes = ledger_to_excel_bytes(ledger_df, subjects)
                st.download_button(
                    "Download Ledger (XLSX)",
                    data=excel_bytes,
                    file_name="srms_final_ledger.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Excel export requires `openpyxl`. Error: {e}")


with tab_print:
    st.markdown("### Marksheet Preview + Bulk PDF")
    ledger_df = st.session_state.srms_data.get("ledger_df")
    subjects = cfg["subjects"]

    if ledger_df is None or ledger_df.empty:
        st.info("Calculate the ledger first from **Data & Ledger**.")
    else:
        col_a, col_b = st.columns([0.7, 1.3], gap="large")
        with col_a:
            st.markdown("#### Preview Options")
            roll_list = ledger_df["Roll_No"].astype(str).tolist()
            pick = st.selectbox("Select Roll No", options=roll_list, index=0)
            row = ledger_df[ledger_df["Roll_No"].astype(str) == str(pick)].iloc[0]
            verification_links = st.session_state.verification_links
            current_verification_url = verification_url_for_row(row, verification_links)

            st.markdown("#### QR Verification")
            if current_verification_url:
                st.success("This marksheet has a published verification PDF.")
                st.link_button("Open Verification PDF", current_verification_url, use_container_width=True)
            if st.button("Publish This Marksheet QR/PDF", use_container_width=True, disabled=not supabase_configured()):
                with st.spinner("Creating PDF, uploading to Supabase, and generating QR link..."):
                    public_url, err = publish_marksheet_pdf(row, cfg, subjects)
                if err:
                    st.error(err)
                elif public_url:
                    st.session_state.verification_links[verification_key(row)] = public_url
                    st.success("Verification PDF published. The QR is now added to this marksheet.")
                    st.rerun()
            if not supabase_configured():
                st.info("Configure Supabase secrets to enable QR/PDF publishing.")

            html = marksheet_html(row, cfg, subjects, current_verification_url)
            st.download_button(
                "Download This Marksheet (HTML)",
                data=html.encode("utf-8"),
                file_name=f"marksheet_{pick}.html",
                mime="text/html",
                use_container_width=True,
            )

            st.markdown("#### Bulk PDF (Class)")
            st.markdown('<div class="srms-muted srms-small">Creates <b>one PDF</b> containing individual marksheets for the entire class (sorted by Roll No).</div>', unsafe_allow_html=True)

            if st.button("Publish QR PDFs For Whole Class", use_container_width=True, disabled=not supabase_configured()):
                progress = st.progress(0)
                published = 0
                first_error = ""
                for idx, (_, student) in enumerate(ledger_df.iterrows(), start=1):
                    public_url, err = publish_marksheet_pdf(student, cfg, subjects)
                    if err and not first_error:
                        first_error = err
                    if public_url:
                        st.session_state.verification_links[verification_key(student)] = public_url
                        published += 1
                    progress.progress(idx / len(ledger_df))
                if first_error:
                    st.error(f"Published {published} marksheet(s), then hit: {first_error}")
                else:
                    st.success(f"Published {published} verification PDF(s).")
                    st.rerun()

            if st.button("Generate Bulk PDF", type="primary", use_container_width=True):
                with st.spinner("Generating PDF..."):
                    pdf_bytes, err = bulk_marksheet_pdf(ledger_df, cfg, subjects, st.session_state.verification_links)
                if err:
                    st.error(err)
                else:
                    st.success("Bulk PDF generated.")
                    st.download_button(
                        "Download Bulk Marksheet PDF",
                        data=pdf_bytes,
                        file_name="srms_bulk_marksheets.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                    )

            st.markdown("#### Print-All (HTML)")
            st.markdown(
                '<div class="srms-muted srms-small">Generates a single long HTML (with page breaks). Open it and press <b>Ctrl + P</b> to print the whole class at once.</div>',
                unsafe_allow_html=True,
            )
            class_html = bulk_marksheet_print_html(ledger_df, cfg, subjects, st.session_state.verification_links)
            st.download_button(
                "Download Print-All HTML",
                data=class_html.encode("utf-8"),
                file_name="srms_print_all_marksheets.html",
                mime="text/html",
                use_container_width=True,
            )

        with col_b:
            st.markdown("#### Marksheet (HTML Preview)")
            st.components.v1.html(html, height=900, scrolling=True)


with tab_insights:
    st.markdown("### Dashboard")
    ledger_df = st.session_state.srms_data.get("ledger_df")
    subjects = cfg["subjects"]

    if ledger_df is None or ledger_df.empty:
        st.info("Calculate the ledger first from **Data & Ledger**.")
    else:
        a, b = st.columns([1, 1], gap="large")

        with a:
            st.markdown("#### Grade Distribution (Final Grade)")
            dist = ledger_df["Final_Grade"].fillna("NG").value_counts().reset_index()
            dist.columns = ["Final_Grade", "Count"]
            fig = px.bar(dist, x="Final_Grade", y="Count", text="Count")
            fig.update_layout(height=420, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig, use_container_width=True)

        with b:
            st.markdown("#### Subject-wise Pass/Fail")
            rows = []
            for s in subjects:
                k = s.key
                status_col = f"{k}__status"
                if status_col not in ledger_df.columns:
                    continue
                vc = ledger_df[status_col].fillna("NG").value_counts()
                rows.append({"Subject": s.name, "Pass": int(vc.get("Pass", 0)), "Fail": int(vc.get("Fail", 0))})
            pf = pd.DataFrame(rows)
            if pf.empty:
                st.info("No subject status columns found yet.")
            else:
                pf_melt = pf.melt(id_vars=["Subject"], value_vars=["Pass", "Fail"], var_name="Status", value_name="Count")
                fig2 = px.bar(pf_melt, x="Subject", y="Count", color="Status", barmode="group", text="Count")
                fig2.update_layout(height=420, margin=dict(l=10, r=10, t=10, b=10))
                st.plotly_chart(fig2, use_container_width=True)

        st.markdown("### Quick Filters")
        c1, c2, c3 = st.columns([1, 1, 1], gap="medium")
        with c1:
            show = st.selectbox("Show Students", options=["All", "Pass", "Fail", "NG"], index=0)
        with c2:
            min_cgpa = float(st.number_input("Min CGPA", min_value=0.0, max_value=4.0, value=0.0, step=0.1))
        with c3:
            max_cgpa = float(st.number_input("Max CGPA", min_value=0.0, max_value=4.0, value=4.0, step=0.1))

        filtered = ledger_df.copy()
        if show != "All":
            filtered = filtered[filtered["Remarks"] == show]
        filtered = filtered[(filtered["CGPA"] >= min_cgpa) & (filtered["CGPA"] <= max_cgpa)]
        st.dataframe(filtered, use_container_width=True, height=420)


st.sidebar.markdown("---")
st.sidebar.caption("SRMS • Streamlit + Pandas + Plotly • PDF via xhtml2pdf/pdfkit")
# FINALLYYYYYYYYYYYYYYYYYYYYYYYYY